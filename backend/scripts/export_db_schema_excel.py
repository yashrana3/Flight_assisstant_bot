#!/usr/bin/env python3
"""
Export PostgreSQL data to Excel: one sheet per table (sheet name = table name),
row 1 = column names, following rows = records.

Loads backend/.env automatically. Run from backend:
  python3 scripts/export_db_schema_excel.py

Optional schema reference sheets:
  python3 scripts/export_db_schema_excel.py --include-schema
"""

from __future__ import annotations

import argparse
import json
import sys
import uuid
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

BACKEND_ROOT = Path(__file__).resolve().parent.parent
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from dotenv import load_dotenv  # noqa: E402

load_dotenv(BACKEND_ROOT / ".env")
load_dotenv()

import models  # noqa: F401, E402 — registers all mappers
from database import BaseChat, BaseUser, engine_chat, engine_user  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from sqlalchemy import text  # noqa: E402
from sqlalchemy.schema import UniqueConstraint  # noqa: E402


def _safe_default(col) -> str:
    d = col.default
    if d is None:
        return ""
    if hasattr(d, "arg"):
        arg = d.arg
        if callable(arg):
            return "(callable default)"
        return str(arg)
    return str(d)


def _fk_refs(table) -> list[str]:
    out: list[str] = []
    for fk in table.foreign_key_constraints:
        targets = [f"{e.column.table.name}.{e.column.name}" for e in fk.elements]
        src = ", ".join(fk.column_keys)
        tgt = ", ".join(targets)
        ondel = fk.ondelete or "NO ACTION"
        out.append(f"{src} → {tgt} ON DELETE {ondel}")
    return out


def _collect_columns(database_label: str, table) -> list[dict]:
    rows: list[dict] = []
    pk_cols = set(table.primary_key.columns.keys())
    indexed = set()
    for ix in table.indexes:
        for c in ix.columns:
            indexed.add(c.name)

    fk_by_col: dict[str, list[str]] = {}
    for fk in table.foreign_key_constraints:
        for col, elem in zip(fk.columns, fk.elements, strict=True):
            fk_by_col.setdefault(col.name, []).append(
                f"{elem.column.table.name}.{elem.column.name}"
            )

    unique_cols: set[str] = set()
    for c in table.constraints:
        if isinstance(c, UniqueConstraint):
            for col in c.columns:
                unique_cols.add(col.name)

    for col in table.columns:
        fk = "; ".join(fk_by_col.get(col.name, []))
        note_parts = []
        if col.name == "metadata" and database_label == "Chat DB":
            note_parts.append('ORM: ChatMessage.metadata_ maps to column "metadata"')
        rows.append(
            {
                "database": database_label,
                "table": table.name,
                "column_name": col.name,
                "sqlalchemy_type": str(col.type),
                "nullable": "YES" if col.nullable else "NO",
                "default": _safe_default(col),
                "primary_key": "YES" if col.name in pk_cols else "",
                "unique": "YES" if col.unique or col.name in unique_cols else "",
                "indexed": "YES" if col.name in indexed else "",
                "foreign_key": fk,
                "notes": "; ".join(note_parts),
            }
        )
    return rows


def _table_notes(table) -> str:
    notes = []
    if table.name == "chat_sessions" and "user_id" in table.columns:
        notes.append("user_id: same value as users.id (User DB); no FK across databases")
    if table.name in ("consent_records", "feedback") and "session_id" in table.columns:
        notes.append("session_id: chat_sessions.id (Chat DB); no FK across databases")
    return " ".join(notes)


def _sanitize_sheet_name(raw: str, used: set[str]) -> str:
    invalid = "[]:*?/\\"
    s = "".join("_" if c in invalid else c for c in raw)
    s = s.strip()[:31] or "table"
    candidate = s
    n = 2
    lower_used = {u.lower() for u in used}
    while candidate.lower() in lower_used:
        suffix = f"_{n}"
        candidate = (s[: 31 - len(suffix)] + suffix)[:31]
        n += 1
    used.add(candidate)
    return candidate


def _cell_value(val: object) -> object:
    if val is None:
        return ""
    if isinstance(val, uuid.UUID):
        return str(val)
    if isinstance(val, (datetime, date)):
        return val.isoformat()
    if isinstance(val, Decimal):
        return str(val)
    if isinstance(val, bytes):
        return val.decode("utf-8", errors="replace")
    if isinstance(val, (list, dict)):
        return json.dumps(val, default=str)
    return val


def _append_data_sheet(
    wb: Workbook,
    engine,
    table,
    sheet_title: str,
    header_fill: PatternFill,
    header_font: Font,
) -> None:
    ws = wb.create_sheet(sheet_title)
    q = text(f'SELECT * FROM "{table.name}"')
    try:
        with engine.connect() as conn:
            result = conn.execute(q)
            keys = list(result.keys())
            if not keys:
                ws.append(["(no columns)"])
                return
            ws.append(keys)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            for row in result:
                ws.append([_cell_value(row[k]) for k in keys])
    except Exception as exc:  # noqa: BLE001 — surface DB errors in-sheet
        ws.append(["error"])
        ws.append([str(exc)])


def _style_header_row(ws, header_fill: PatternFill, header_font: Font) -> None:
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _add_schema_sheets(wb: Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)

    ws_tables = wb.create_sheet("Schema_Tables")
    ws_tables.append(
        ["database", "table_name", "model_class", "foreign_keys", "cross_db_notes"]
    )
    _style_header_row(ws_tables, header_fill, header_font)

    mapper_by_table: dict[str, str] = {}
    for base in (BaseUser, BaseChat):
        for m in base.registry.mappers:
            t = getattr(m.class_, "__table__", None)
            if t is not None:
                mapper_by_table[t.name] = m.class_.__name__

    def add_table_rows(metadata, label: str) -> None:
        for tbl in metadata.sorted_tables:
            fks = "; ".join(_fk_refs(tbl)) or ""
            ws_tables.append(
                [
                    label,
                    tbl.name,
                    mapper_by_table.get(tbl.name, ""),
                    fks,
                    _table_notes(tbl),
                ]
            )

    add_table_rows(BaseUser.metadata, "User DB")
    add_table_rows(BaseChat.metadata, "Chat DB")

    for col in ("A", "B", "C", "D", "E"):
        ws_tables.column_dimensions[col].width = 28 if col != "E" else 40

    ws_cols = wb.create_sheet("Schema_Columns")
    ws_cols.append(
        [
            "database",
            "table_name",
            "column_name",
            "sqlalchemy_type",
            "nullable",
            "default",
            "primary_key",
            "unique",
            "indexed",
            "foreign_key",
            "notes",
        ]
    )
    _style_header_row(ws_cols, header_fill, header_font)

    for metadata, label in ((BaseUser.metadata, "User DB"), (BaseChat.metadata, "Chat DB")):
        for tbl in metadata.sorted_tables:
            for row in _collect_columns(label, tbl):
                ws_cols.append([row[k] for k in row])

    widths = {
        "A": 12,
        "B": 22,
        "C": 22,
        "D": 28,
        "E": 10,
        "F": 24,
        "G": 14,
        "H": 10,
        "I": 10,
        "J": 36,
        "K": 36,
    }
    for letter, w in widths.items():
        ws_cols.column_dimensions[letter].width = w


def build_workbook(*, include_schema: bool) -> Workbook:
    wb = Workbook()
    ws_readme = wb.active
    ws_readme.title = "Readme"
    ws_readme["A1"] = "BookwithAI — database export (one sheet per table)"
    ws_readme["A1"].font = Font(bold=True, size=14)
    ws_readme["A3"] = (
        "Each table sheet: row 1 = column names, next rows = current database rows. "
        "User DB tables use USER_DATABASE_URL; Chat DB tables use CHAT_DATABASE_URL "
        "(often the same database)."
    )
    ws_readme["A5"] = (
        "Regenerate: python3 backend/scripts/export_db_schema_excel.py "
        "(from repo root) or python3 scripts/export_db_schema_excel.py (from backend)."
    )
    ws_readme.column_dimensions["A"].width = 100

    header_fill = PatternFill("solid", fgColor="2E7D32")
    header_font = Font(color="FFFFFF", bold=True)

    used_sheet_names: set[str] = {"Readme"}

    for tbl in BaseUser.metadata.sorted_tables:
        title = _sanitize_sheet_name(tbl.name, used_sheet_names)
        _append_data_sheet(wb, engine_user, tbl, title, header_fill, header_font)

    for tbl in BaseChat.metadata.sorted_tables:
        title = _sanitize_sheet_name(tbl.name, used_sheet_names)
        _append_data_sheet(wb, engine_chat, tbl, title, header_fill, header_font)

    if include_schema:
        _add_schema_sheets(wb)

    return wb


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Export DB tables to Excel (sheet per table + rows)."
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=BACKEND_ROOT.parent / "docs" / "database_export.xlsx",
        help="Output .xlsx path",
    )
    parser.add_argument(
        "--include-schema",
        action="store_true",
        help="Add Schema_Tables and Schema_Columns reference sheets",
    )
    args = parser.parse_args()
    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook(include_schema=args.include_schema)
    wb.save(args.output)
    print(f"Wrote {args.output}")


if __name__ == "__main__":
    main()
